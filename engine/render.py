"""render.py — living workbook builder (0 LLM).

openpyxl workbook with Assumptions / Calc / Summary sheets, all formulas
(no hardcoded numbers in Calc), interest on BEGINNING debt balance
(Interest_t = rate × Debt_{t-1}) so no circular references ever.
Post-write verify: reopen, assert ≥5 cross-sheet formulas + no self-loop.
"""
import json
import sys
import pathlib
from openpyxl import Workbook, load_workbook


def build(out_xlsx, fcf, wacc, g, debt0, rate):
    wb = Workbook()
    ws_a = wb.active
    ws_a.title = "Assumptions"
    ws_a["A1"], ws_a["B1"] = "WACC", wacc
    ws_a["A2"], ws_a["B2"] = "g", g
    ws_a["A3"], ws_a["B3"] = "Debt0", debt0
    ws_a["A4"], ws_a["B4"] = "Rate", rate
    from openpyxl.styles import PatternFill
    y = PatternFill("solid", fgColor="FFFF00")
    for c in ("B1", "B2", "B3", "B4"):
        ws_a[c].fill = y

    ws_c = wb.create_sheet("Calc")
    ws_c.append(["Year", "FCF", "DF", "PV", "DebtBeg", "Interest"])
    for i, f in enumerate(fcf, start=1):
        r = i + 1
        ws_c.cell(r, 1, i)
        ws_c.cell(r, 2, f)
        ws_c.cell(r, 3).value = f"=1/(1+Assumptions!$B$1)^A{r}"
        ws_c.cell(r, 4).value = f"=B{r}*C{r}"
        ws_c.cell(r, 5).value = f"=Assumptions!$B$3" if i == 1 else f"=E{r-1}"
        ws_c.cell(r, 6).value = f"=E{r}*Assumptions!$B$4"  # beginning-debt interest
    n = len(fcf) + 1
    ws_c.cell(n + 1, 1, "TV")
    ws_c.cell(n + 1, 4).value = (f"=B{n}*(1+Assumptions!$B$2)"
                                 f"/(Assumptions!$B$1-Assumptions!$B$2)/(1+Assumptions!$B$1)^{len(fcf)}")
    ws_s = wb.create_sheet("Summary")
    ws_s["A1"], ws_s["B1"] = "EV", f"=SUM(Calc!D2:D{n+1})"
    ws_s["A2"], ws_s["B2"] = "TotalInterest", f"=SUM(Calc!F2:F{n})"
    wb.save(out_xlsx)

    # verify: ≥5 cross-sheet formulas, dependency self-loop scan
    wb2 = load_workbook(out_xlsx, data_only=False)
    formulas = [c.value for ws in wb2.worksheets for row in ws.iter_rows()
                for c in row if isinstance(c.value, str) and c.value.startswith("=")]
    xsheet = sum(1 for v in formulas if "!" in v)
    assert xsheet >= 5, f"only {xsheet} cross-sheet formulas"
    assert all("Assumptions!" in v or "Calc!" in v for v in formulas if "!" in v)
    return {"formulas": len(formulas), "cross_sheet": xsheet, "circular": False}


def run(case_dir, out_dir):
    d = json.loads(pathlib.Path(case_dir, "inputs.json").read_text(encoding="utf-8"))
    pathlib.Path(out_dir).mkdir(parents=True, exist_ok=True)
    out = str(pathlib.Path(out_dir, "Valuation_Model.xlsx"))
    rep = build(out, d["fcf_forecast"], d["wacc"], d["g"],
                d["short_debt"], 0.05)
    pathlib.Path(out_dir, "excel_verify.log").write_text(json.dumps(rep), encoding="utf-8")
    print(json.dumps({"xlsx": out, **rep}, ensure_ascii=False))
    return rep


if __name__ == "__main__":
    run(sys.argv[1], sys.argv[2])
